# SAP FI - Accounts Payable Processing & Aging Report
# Simulates: FB60 (Vendor Invoice), F110 (Auto Payment), FBL1N (Vendor Line Items)
# Company Code: IN01 | Fiscal Year: 2024 | Period: 1 (April 2024)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

REPORT_DATE = datetime(2024, 5, 31)

print("=" * 60)
print("  SAP FI - ACCOUNTS PAYABLE REPORT")
print("  Company Code: IN01 | Period: April 2024")
print("=" * 60)

# ── Load Data ──────────────────────────────────────────────────
ap = pd.read_csv("transactions/vendor_invoices.csv")
vm = pd.read_csv("master_data/vendor_master.csv")
ap["Posting_Date"] = pd.to_datetime(ap["Posting_Date"])
ap["Due_Date"]     = pd.to_datetime(ap["Due_Date"])

# ── Step 1: Vendor Invoice Summary (FB60) ─────────────────────
print("\n[1] Vendor Invoice Summary (T-code: FB60)...")
summary = ap.groupby(["Vendor_ID","Vendor_Name"]).agg(
    Total_Invoices=("Doc_Number","count"),
    Total_Invoice_Amount=("Invoice_Amount","sum"),
    Total_Paid=("Payment_Amount","sum"),
    Total_Outstanding=("Outstanding_Amount","sum")
).reset_index()
print(summary.to_string(index=False))

# ── Step 2: AP Aging Report (FBL1N) ───────────────────────────
print("\n[2] AP Aging Report (T-code: FBL1N)...")
open_items = ap[ap["Outstanding_Amount"] > 0].copy()
open_items["Days_Overdue"] = (REPORT_DATE - open_items["Due_Date"]).dt.days

def aging_bucket(days):
    if days <= 0:   return "Not Yet Due"
    elif days <= 30: return "1-30 Days"
    elif days <= 60: return "31-60 Days"
    elif days <= 90: return "61-90 Days"
    else:            return "90+ Days"

open_items["Aging_Bucket"] = open_items["Days_Overdue"].apply(aging_bucket)

aging_summary = open_items.groupby("Aging_Bucket")["Outstanding_Amount"].sum().reset_index()
aging_summary.columns = ["Aging Bucket","Outstanding Amount (INR)"]
print(aging_summary.to_string(index=False))
print(f"\n    Total AP Outstanding : INR {open_items['Outstanding_Amount'].sum():,.2f}")

# ── Step 3: Payment Run Summary (F110) ────────────────────────
print("\n[3] Payment Run Summary (T-code: F110)...")
paid = ap[ap["Payment_Amount"] > 0][["Doc_Number","Vendor_Name","Invoice_Amount","Payment_Amount","Payment_Date","Payment_Doc"]]
print(paid.to_string(index=False))
print(f"\n    Total Payments Made : INR {paid['Payment_Amount'].sum():,.2f}")

# ── Step 4: Export to Excel ────────────────────────────────────
print("\n[4] Exporting AP Report to Excel...")

wb = Workbook()

# ── Sheet 1: Vendor Invoice List ──────────────────────────────
ws1 = wb.active
ws1.title = "Vendor Invoices (FB60)"

header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill  = PatternFill("solid", fgColor="833C00")
normal_font  = Font(name="Calibri", size=10)
total_font   = Font(name="Calibri", bold=True, size=11)
total_fill   = PatternFill("solid", fgColor="FCE4D6")
center_align = Alignment(horizontal="center")
right_align  = Alignment(horizontal="right")
thin_border  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

ws1.merge_cells("A1:H1")
ws1["A1"] = "PrecisionParts Pvt. Ltd. — Vendor Invoice Report (FB60)"
ws1["A1"].font = Font(name="Calibri", bold=True, size=13, color="833C00")
ws1["A1"].alignment = center_align

ws1.merge_cells("A2:H2")
ws1["A2"] = f"Company Code: IN01 | Report Date: {REPORT_DATE.strftime('%d-%b-%Y')} | T-Code: FB60 / FBL1N"
ws1["A2"].font = Font(name="Calibri", size=10, italic=True)
ws1["A2"].alignment = center_align

ws1.append([])

headers = ["Doc No","Posting Date","Due Date","Vendor ID",
           "Vendor Name","Invoice Amt (INR)","Paid Amt (INR)","Outstanding (INR)"]
ws1.append(headers)
for c in range(1, 9):
    cell = ws1.cell(row=4, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for _, row in ap.iterrows():
    ws1.append([
        row["Doc_Number"],
        row["Posting_Date"].strftime("%d-%b-%Y"),
        row["Due_Date"].strftime("%d-%b-%Y"),
        row["Vendor_ID"],
        row["Vendor_Name"],
        row["Invoice_Amount"],
        row["Payment_Amount"],
        row["Outstanding_Amount"]
    ])
    r = ws1.max_row
    for c in range(1, 9):
        cell = ws1.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        if c in [6, 7, 8]:
            cell.number_format = '#,##0.00'
            cell.alignment = right_align

total_row = ws1.max_row + 1
ws1.append(["","","","","TOTAL",
            ap["Invoice_Amount"].sum(),
            ap["Payment_Amount"].sum(),
            ap["Outstanding_Amount"].sum()])
for c in range(1, 9):
    cell = ws1.cell(row=total_row, column=c)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    if c in [6, 7, 8]:
        cell.number_format = '#,##0.00'
        cell.alignment = right_align

col_widths = [10, 15, 15, 12, 28, 20, 18, 20]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: AP Aging Report ──────────────────────────────────
ws2 = wb.create_sheet("AP Aging (FBL1N)")

aging_header_fill = PatternFill("solid", fgColor="375623")
bucket_fills = {
    "Not Yet Due":  PatternFill("solid", fgColor="E2EFDA"),
    "1-30 Days":    PatternFill("solid", fgColor="FFEB9C"),
    "31-60 Days":   PatternFill("solid", fgColor="FFCC00"),
    "61-90 Days":   PatternFill("solid", fgColor="FF9900"),
    "90+ Days":     PatternFill("solid", fgColor="FF0000"),
}

ws2.merge_cells("A1:F1")
ws2["A1"] = "PrecisionParts Pvt. Ltd. — AP Aging Report (FBL1N)"
ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color="375623")
ws2["A1"].alignment = center_align

ws2.merge_cells("A2:F2")
ws2["A2"] = f"Company Code: IN01 | Aging As Of: {REPORT_DATE.strftime('%d-%b-%Y')} | T-Code: FBL1N"
ws2["A2"].font = Font(name="Calibri", size=10, italic=True)
ws2["A2"].alignment = center_align
ws2.append([])

aging_headers = ["Doc No","Vendor Name","Due Date","Outstanding (INR)","Days Overdue","Aging Bucket"]
ws2.append(aging_headers)
for c in range(1, 7):
    cell = ws2.cell(row=4, column=c)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="375623")
    cell.alignment = center_align
    cell.border = thin_border

for _, row in open_items.iterrows():
    ws2.append([
        row["Doc_Number"],
        row["Vendor_Name"],
        row["Due_Date"].strftime("%d-%b-%Y"),
        row["Outstanding_Amount"],
        int(row["Days_Overdue"]),
        row["Aging_Bucket"]
    ])
    r = ws2.max_row
    bucket = row["Aging_Bucket"]
    for c in range(1, 7):
        cell = ws2.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = bucket_fills.get(bucket, PatternFill())
        if c == 4:
            cell.number_format = '#,##0.00'
            cell.alignment = right_align

col_widths2 = [10, 28, 15, 20, 15, 15]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save("reports/ap_aging_report.xlsx")
print("    ✔ Saved → reports/ap_aging_report.xlsx")
print("\n" + "=" * 60)
print("  AP PROCESSING SIMULATION COMPLETE")
print("=" * 60)
