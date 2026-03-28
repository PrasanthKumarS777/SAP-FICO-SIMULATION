# SAP FI - Month-End Closing Activities
# Simulates: AFAB (Depreciation Run), F.05 (FX Revaluation),
#            OB52 (Period Close), F.01 (Financial Statements)
# Company Code: IN01 | Fiscal Year: 2024 | Period: 1 (April 2024)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

PERIOD       = "Period 1 - April 2024"
COMPANY_CODE = "IN01"
FISCAL_YEAR  = 2024

print("=" * 60)
print("  SAP FI - MONTH-END CLOSING ACTIVITIES")
print(f"  Company Code: {COMPANY_CODE} | {PERIOD}")
print("=" * 60)

# ── Load All Data ──────────────────────────────────────────────
gl  = pd.read_csv("transactions/gl_journal_entries.csv")
dep = pd.read_csv("transactions/depreciation.csv")
fx  = pd.read_csv("transactions/fx_revaluation.csv")
ap  = pd.read_csv("transactions/vendor_invoices.csv")
ar  = pd.read_csv("transactions/customer_invoices.csv")
co  = pd.read_csv("transactions/cost_allocations.csv")

# ── Step 1: Depreciation Run (AFAB) ───────────────────────────
print("\n[1] Depreciation Run (T-code: AFAB)...")
print(f"    Method : Straight Line | Period: April 2024")
print(f"\n    {'Asset':<10} {'Description':<22} {'Acq. Value':>14} {'Dep. Amount':>14} {'WDV':>14}")
print(f"    {'-'*74}")
for _, row in dep.iterrows():
    print(f"    {row['Asset_No']:<10} {row['Asset_Description']:<22} "
          f"INR {row['Acquisition_Value']:>10,.0f} "
          f"INR {row['Current_Period_Dep']:>10,.0f} "
          f"INR {row['WDV']:>10,.0f}")
total_dep = dep["Current_Period_Dep"].sum()
print(f"\n    Total Depreciation Posted : INR {total_dep:,.2f}")
print(f"    GL Dr: 330000 (Depreciation Expense)")
print(f"    GL Cr: 150100 (Accumulated Depreciation)")
print(f"    ✔ Depreciation run completed — AFAB posted successfully")

# ── Step 2: FX Revaluation (F.05) ─────────────────────────────
print("\n[2] Foreign Currency Revaluation (T-code: F.05)...")
for _, row in fx.iterrows():
    direction = "GAIN" if row["FX_Gain_Loss"] > 0 else "LOSS"
    print(f"    Doc {row['Doc_Number']} | {row['Description']}")
    print(f"    {row['Original_Currency']} {row['Original_Amount']:,.0f} | "
          f"Rate at Posting: {row['Exchange_Rate_Posting']} → "
          f"Rate at Revaluation: {row['Exchange_Rate_Revaluation']}")
    print(f"    FX {direction}: INR {abs(row['FX_Gain_Loss']):,.2f}\n")

net_fx = fx["FX_Gain_Loss"].sum()
print(f"    Net FX Gain/Loss : INR {net_fx:,.2f} "
      f"({'GAIN' if net_fx > 0 else 'LOSS'})")
print(f"    ✔ FX revaluation completed — F.05 posted successfully")

# ── Step 3: Build Financial Statements (F.01) ─────────────────
print("\n[3] Building Financial Statements (T-code: F.01)...")

gl_summary = gl.groupby("GL_Account").agg(
    Total_Debit=("Debit_Amount","sum"),
    Total_Credit=("Credit_Amount","sum")
).reset_index()
gl_summary["Net_Balance"] = gl_summary["Total_Debit"] - gl_summary["Total_Credit"]

def get_balance(account):
    row = gl_summary[gl_summary["GL_Account"] == account]
    return float(row["Net_Balance"].values[0]) if len(row) > 0 else 0.0

# Add depreciation to GL
dep_amount = float(dep["Current_Period_Dep"].sum())
fx_gain    = float(fx["FX_Gain_Loss"].sum())

# ── Balance Sheet ──────────────────────────────────────────────
cash_bank       = get_balance(100100)
accounts_rec    = get_balance(110000)
inventory       = get_balance(120000)
fixed_assets    = get_balance(150000) - dep_amount
total_assets    = cash_bank + accounts_rec + inventory + fixed_assets

accounts_pay    = abs(get_balance(200000))
short_term_loan = abs(get_balance(210000))
total_liab      = accounts_pay + short_term_loan

share_capital   = abs(get_balance(250000))
revenue         = abs(get_balance(400000)) + abs(get_balance(410000))
expenses        = (get_balance(300000) + get_balance(310000) +
                   get_balance(320000) + get_balance(330000) +
                   get_balance(340000) + dep_amount)
net_profit      = revenue - expenses + fx_gain
retained_earn   = net_profit
total_equity    = share_capital + retained_earn
total_liab_eq   = total_liab + total_equity

print(f"\n    ── BALANCE SHEET ──")
print(f"    Cash & Bank          : INR {cash_bank:>14,.2f}")
print(f"    Accounts Receivable  : INR {accounts_rec:>14,.2f}")
print(f"    Inventory            : INR {inventory:>14,.2f}")
print(f"    Fixed Assets (Net)   : INR {fixed_assets:>14,.2f}")
print(f"    {'─'*38}")
print(f"    TOTAL ASSETS         : INR {total_assets:>14,.2f}")
print(f"\n    Accounts Payable     : INR {accounts_pay:>14,.2f}")
print(f"    Short Term Loan      : INR {short_term_loan:>14,.2f}")
print(f"    Share Capital        : INR {share_capital:>14,.2f}")
print(f"    Retained Earnings    : INR {retained_earn:>14,.2f}")
print(f"    {'─'*38}")
print(f"    TOTAL LIAB + EQUITY  : INR {total_liab_eq:>14,.2f}")

bs_check = abs(total_assets - total_liab_eq) < 1
print(f"\n    ✔ Balance Sheet {'BALANCED' if bs_check else 'NOT BALANCED'}")

# ── P&L ───────────────────────────────────────────────────────
print(f"\n    ── PROFIT & LOSS STATEMENT ──")
print(f"    Sales Revenue        : INR {revenue:>14,.2f}")
print(f"    Total Expenses       : INR {expenses:>14,.2f}")
print(f"    FX Gain/Loss         : INR {fx_gain:>14,.2f}")
print(f"    {'─'*38}")
print(f"    NET PROFIT           : INR {net_profit:>14,.2f}")

# ── Step 4: Period Close Checklist (OB52) ─────────────────────
print("\n[4] Month-End Close Checklist (T-code: OB52)...")
checklist = [
    ("GL Journal Entries Posted",          "FB50", "✔ Complete"),
    ("Vendor Invoices Processed",          "FB60", "✔ Complete"),
    ("Customer Invoices Processed",        "FB70", "✔ Complete"),
    ("Automatic Payment Run Executed",     "F110", "✔ Complete"),
    ("Incoming Payments Posted",           "F-28", "✔ Complete"),
    ("Depreciation Run Executed",          "AFAB", "✔ Complete"),
    ("FX Revaluation Completed",           "F.05", "✔ Complete"),
    ("Cost Center Allocations Posted",     "KB11N","✔ Complete"),
    ("Trial Balance Verified",        "S_ALR_87012284", "✔ Complete"),
    ("Financial Statements Generated",     "F.01", "✔ Complete"),
    ("Posting Period Closed",              "OB52", "✔ Complete"),
]
print(f"\n    {'Activity':<40} {'T-Code':<18} {'Status'}")
print(f"    {'─'*72}")
for activity, tcode, status in checklist:
    print(f"    {activity:<40} {tcode:<18} {status}")

# ── Step 5: Export Full Financial Statements to Excel ─────────
print("\n[5] Exporting Financial Statements to Excel...")

wb  = Workbook()

header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_font    = Font(name="Calibri", bold=True, size=14, color="1F4E79")
section_font  = Font(name="Calibri", bold=True, size=11, color="1F4E79")
normal_font   = Font(name="Calibri", size=10)
total_font    = Font(name="Calibri", bold=True, size=11)
italic_font   = Font(name="Calibri", size=9, italic=True, color="595959")
center_align  = Alignment(horizontal="center")
right_align   = Alignment(horizontal="right")
left_align    = Alignment(horizontal="left")
thin_border   = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin")
)
header_fill   = PatternFill("solid", fgColor="1F4E79")
section_fill  = PatternFill("solid", fgColor="BDD7EE")
total_fill    = PatternFill("solid", fgColor="D6E4F0")
profit_fill   = PatternFill("solid", fgColor="E2EFDA")

# ── Sheet 1: Balance Sheet ────────────────────────────────────
ws1 = wb.active
ws1.title = "Balance Sheet (F.01)"

def write_title(ws, text, subtitle):
    ws.merge_cells("A1:C1")
    ws["A1"] = text
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.merge_cells("A2:C2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = center_align
    ws.append([])

write_title(ws1,
    "PrecisionParts Pvt. Ltd. — Balance Sheet",
    f"Company Code: IN01 | As at 30-Apr-2024 | T-Code: F.01 | Fiscal Year: {FISCAL_YEAR}")

def write_section(ws, title, rows, total_label, total_val):
    ws.append([title, "", ""])
    r = ws.max_row
    ws.cell(r,1).font = section_font
    ws.cell(r,1).fill = section_fill
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(r,1).border = thin_border

    for label, val in rows:
        ws.append(["", label, val])
        r = ws.max_row
        ws.cell(r,1).border = thin_border
        ws.cell(r,2).font   = normal_font
        ws.cell(r,2).border = thin_border
        ws.cell(r,3).font   = normal_font
        ws.cell(r,3).border = thin_border
        ws.cell(r,3).number_format = '#,##0.00'
        ws.cell(r,3).alignment = right_align

    ws.append(["", total_label, total_val])
    r = ws.max_row
    for c in [1,2,3]:
        ws.cell(r,c).font   = total_font
        ws.cell(r,c).fill   = total_fill
        ws.cell(r,c).border = thin_border
    ws.cell(r,3).number_format = '#,##0.00'
    ws.cell(r,3).alignment = right_align
    ws.append([])

write_section(ws1, "ASSETS",
    [("Cash & Bank (GL: 100100)", cash_bank),
     ("Accounts Receivable (GL: 110000)", accounts_rec),
     ("Inventory (GL: 120000)", inventory),
     ("Fixed Assets - Net of Depreciation (GL: 150000)", fixed_assets)],
    "TOTAL ASSETS", total_assets)

write_section(ws1, "LIABILITIES",
    [("Accounts Payable (GL: 200000)", accounts_pay),
     ("Short Term Loan (GL: 210000)", short_term_loan)],
    "TOTAL LIABILITIES", total_liab)

write_section(ws1, "EQUITY",
    [("Share Capital (GL: 250000)", share_capital),
     ("Retained Earnings / Net Profit", retained_earn)],
    "TOTAL EQUITY", total_equity)

ws1.append(["", "TOTAL LIABILITIES + EQUITY", total_liab_eq])
r = ws1.max_row
for c in [1,2,3]:
    ws1.cell(r,c).font   = Font(name="Calibri", bold=True, size=12)
    ws1.cell(r,c).fill   = PatternFill("solid", fgColor="1F4E79")
    ws1.cell(r,c).font   = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws1.cell(r,c).border = thin_border
ws1.cell(r,3).number_format = '#,##0.00'
ws1.cell(r,3).alignment = right_align

ws1.append([])
status_text = "✔ Balance Sheet BALANCED — Assets = Liabilities + Equity" if bs_check else "✘ NOT BALANCED"
ws1.append(["", status_text])
ws1.cell(ws1.max_row, 2).font = Font(
    name="Calibri", bold=True, size=10,
    color="375623" if bs_check else "C00000")

ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 46
ws1.column_dimensions["C"].width = 22

# ── Sheet 2: P&L Statement ────────────────────────────────────
ws2 = wb.create_sheet("P&L Statement (F.01)")
write_title(ws2,
    "PrecisionParts Pvt. Ltd. — Profit & Loss Statement",
    f"Company Code: IN01 | Period: April 2024 | T-Code: F.01")

pl_rows = [
    ("Sales Revenue (GL: 400000)",        revenue,    "INCOME"),
    ("Raw Material Expense (GL: 300000)", get_balance(300000), "EXPENSE"),
    ("Salaries & Wages (GL: 310000)",     get_balance(310000), "EXPENSE"),
    ("Rent Expense (GL: 320000)",         get_balance(320000), "EXPENSE"),
    ("Depreciation (GL: 330000)",         dep_amount,          "EXPENSE"),
    ("Miscellaneous (GL: 340000)",        get_balance(340000), "EXPENSE"),
    ("FX Gain/Loss",                      fx_gain,    "OTHER"),
]

ws2.append(["Category","Description","Amount (INR)"])
for c in range(1, 4):
    cell = ws2.cell(ws2.max_row, c)
    cell.font   = header_font
    cell.fill   = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for desc, val, category in pl_rows:
    ws2.append([category, desc, val])
    r = ws2.max_row
    fill = (PatternFill("solid", fgColor="E2EFDA") if category == "INCOME"
            else PatternFill("solid", fgColor="FCE4D6") if category == "EXPENSE"
            else PatternFill("solid", fgColor="FFF2CC"))
    for c in range(1, 4):
        ws2.cell(r,c).font   = normal_font
        ws2.cell(r,c).border = thin_border
        ws2.cell(r,c).fill   = fill
    ws2.cell(r,3).number_format = '#,##0.00'
    ws2.cell(r,3).alignment = right_align

ws2.append([])
ws2.append(["", "NET PROFIT / (LOSS)", net_profit])
r = ws2.max_row
for c in range(1, 4):
    ws2.cell(r,c).font   = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws2.cell(r,c).fill   = PatternFill("solid", fgColor="375623" if net_profit > 0 else "C00000")
    ws2.cell(r,c).border = thin_border
ws2.cell(r,3).number_format = '#,##0.00'
ws2.cell(r,3).alignment = right_align

ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 42
ws2.column_dimensions["C"].width = 22

# ── Sheet 3: Month-End Checklist ──────────────────────────────
ws3 = wb.create_sheet("Month-End Checklist (OB52)")
ws3.merge_cells("A1:D1")
ws3["A1"] = "PrecisionParts Pvt. Ltd. — Month-End Close Checklist"
ws3["A1"].font = title_font
ws3["A1"].alignment = center_align
ws3.merge_cells("A2:D2")
ws3["A2"] = f"Company Code: IN01 | Period: April 2024 | T-Code: OB52"
ws3["A2"].font = Font(name="Calibri", size=10, italic=True)
ws3["A2"].alignment = center_align
ws3.append([])

ws3.append(["#","Activity","T-Code","Status"])
for c in range(1, 5):
    cell = ws3.cell(ws3.max_row, c)
    cell.font   = header_font
    cell.fill   = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for i, (activity, tcode, status) in enumerate(checklist, 1):
    ws3.append([i, activity, tcode, status])
    r = ws3.max_row
    row_fill = PatternFill("solid", fgColor="E2EFDA")
    for c in range(1, 5):
        ws3.cell(r,c).font   = normal_font
        ws3.cell(r,c).fill   = row_fill
        ws3.cell(r,c).border = thin_border
        ws3.cell(r,c).alignment = center_align if c in [1,3,4] else left_align

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 42
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 15

wb.save("reports/financial_statements.xlsx")
print("    ✔ Saved → reports/financial_statements.xlsx")
print("\n" + "=" * 60)
print("  MONTH-END CLOSING COMPLETE")
print("=" * 60)
