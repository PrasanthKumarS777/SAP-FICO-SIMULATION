# SAP FI - Accounts Receivable Processing & Aging Report
# Simulates: FB70 (Customer Invoice), F-28 (Cash Receipt), FBL5N (Customer Line Items)
# Company Code: IN01 | Fiscal Year: 2024 | Period: 1 (April 2024)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

REPORT_DATE = datetime(2024, 5, 31)

print("=" * 60)
print("  SAP FI - ACCOUNTS RECEIVABLE REPORT")
print("  Company Code: IN01 | Period: April 2024")
print("=" * 60)

# ── Load Data ──────────────────────────────────────────────────
ar = pd.read_csv("transactions/customer_invoices.csv")
ar["Posting_Date"] = pd.to_datetime(ar["Posting_Date"])
ar["Due_Date"]     = pd.to_datetime(ar["Due_Date"])

# ── Step 1: Customer Invoice Summary (FB70) ───────────────────
print("\n[1] Customer Invoice Summary (T-code: FB70)...")
summary = ar.groupby(["Customer_ID","Customer_Name"]).agg(
    Total_Invoices=("Doc_Number","count"),
    Total_Invoice_Amount=("Invoice_Amount","sum"),
    Total_Received=("Payment_Amount","sum"),
    Total_Outstanding=("Outstanding_Amount","sum")
).reset_index()
print(summary.to_string(index=False))

# ── Step 2: Cash Receipt Summary (F-28) ───────────────────────
print("\n[2] Cash Receipt Summary (T-code: F-28)...")
received = ar[ar["Payment_Amount"] > 0][["Doc_Number","Customer_Name","Invoice_Amount","Payment_Amount","Payment_Date","Payment_Doc"]]
print(received.to_string(index=False))
print(f"\n    Total Receipts Posted : INR {received['Payment_Amount'].sum():,.2f}")

# ── Step 3: AR Aging Report (FBL5N) ───────────────────────────
print("\n[3] AR Aging Report (T-code: FBL5N)...")
open_items = ar[ar["Outstanding_Amount"] > 0].copy()
open_items["Days_Overdue"] = (REPORT_DATE - open_items["Due_Date"]).dt.days

def aging_bucket(days):
    if days <= 0:    return "Not Yet Due"
    elif days <= 30: return "1-30 Days"
    elif days <= 60: return "31-60 Days"
    elif days <= 90: return "61-90 Days"
    else:            return "90+ Days"

open_items["Aging_Bucket"] = open_items["Days_Overdue"].apply(aging_bucket)

aging_summary = open_items.groupby("Aging_Bucket")["Outstanding_Amount"].sum().reset_index()
aging_summary.columns = ["Aging Bucket", "Outstanding Amount (INR)"]
print(aging_summary.to_string(index=False))
print(f"\n    Total AR Outstanding : INR {open_items['Outstanding_Amount'].sum():,.2f}")

# ── Step 4: Collection Efficiency ─────────────────────────────
print("\n[4] Collection Efficiency Analysis...")
total_invoiced  = ar["Invoice_Amount"].sum()
total_collected = ar["Payment_Amount"].sum()
collection_rate = (total_collected / total_invoiced) * 100
print(f"    Total Invoiced   : INR {total_invoiced:,.2f}")
print(f"    Total Collected  : INR {total_collected:,.2f}")
print(f"    Collection Rate  : {collection_rate:.1f}%")

# ── Step 5: Export to Excel ────────────────────────────────────
print("\n[5] Exporting AR Report to Excel...")

wb = Workbook()

header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
normal_font  = Font(name="Calibri", size=10)
total_font   = Font(name="Calibri", bold=True, size=11)
total_fill   = PatternFill("solid", fgColor="DEEAF1")
center_align = Alignment(horizontal="center")
right_align  = Alignment(horizontal="right")
thin_border  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

# ── Sheet 1: Customer Invoice List ────────────────────────────
ws1 = wb.active
ws1.title = "Customer Invoices (FB70)"

ws1.merge_cells("A1:H1")
ws1["A1"] = "PrecisionParts Pvt. Ltd. — Customer Invoice Report (FB70)"
ws1["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
ws1["A1"].alignment = center_align

ws1.merge_cells("A2:H2")
ws1["A2"] = f"Company Code: IN01 | Report Date: {REPORT_DATE.strftime('%d-%b-%Y')} | T-Code: FB70 / FBL5N"
ws1["A2"].font = Font(name="Calibri", size=10, italic=True)
ws1["A2"].alignment = center_align
ws1.append([])

headers = ["Doc No","Posting Date","Due Date","Customer ID",
           "Customer Name","Invoice Amt (INR)","Received (INR)","Outstanding (INR)"]
ws1.append(headers)
for c in range(1, 9):
    cell = ws1.cell(row=4, column=c)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = center_align
    cell.border = thin_border

for _, row in ar.iterrows():
    ws1.append([
        row["Doc_Number"],
        row["Posting_Date"].strftime("%d-%b-%Y"),
        row["Due_Date"].strftime("%d-%b-%Y"),
        row["Customer_ID"],
        row["Customer_Name"],
        row["Invoice_Amount"],
        row["Payment_Amount"],
        row["Outstanding_Amount"]
    ])
    r = ws1.max_row
    for c in range(1, 9):
        cell = ws1.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        if c in [6, 7, 8]:
            cell.number_format = '#,##0.00'
            cell.alignment = right_align

total_row = ws1.max_row + 1
ws1.append(["","","","","TOTAL",
            ar["Invoice_Amount"].sum(),
            ar["Payment_Amount"].sum(),
            ar["Outstanding_Amount"].sum()])
for c in range(1, 9):
    cell = ws1.cell(row=total_row, column=c)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    if c in [6, 7, 8]:
        cell.number_format = '#,##0.00'
        cell.alignment = right_align

col_widths = [10, 15, 15, 13, 28, 20, 18, 20]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: AR Aging Report ──────────────────────────────────
ws2 = wb.create_sheet("AR Aging (FBL5N)")

bucket_fills = {
    "Not Yet Due":  PatternFill("solid", fgColor="E2EFDA"),
    "1-30 Days":    PatternFill("solid", fgColor="FFEB9C"),
    "31-60 Days":   PatternFill("solid", fgColor="FFCC00"),
    "61-90 Days":   PatternFill("solid", fgColor="FF9900"),
    "90+ Days":     PatternFill("solid", fgColor="FF0000"),
}

ws2.merge_cells("A1:F1")
ws2["A1"] = "PrecisionParts Pvt. Ltd. — AR Aging Report (FBL5N)"
ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
ws2["A1"].alignment = center_align

ws2.merge_cells("A2:F2")
ws2["A2"] = f"Company Code: IN01 | Aging As Of: {REPORT_DATE.strftime('%d-%b-%Y')} | T-Code: FBL5N"
ws2["A2"].font = Font(name="Calibri", size=10, italic=True)
ws2["A2"].alignment = center_align
ws2.append([])

aging_headers = ["Doc No","Customer Name","Due Date","Outstanding (INR)","Days Overdue","Aging Bucket"]
ws2.append(aging_headers)
for c in range(1, 7):
    cell = ws2.cell(row=4, column=c)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = center_align
    cell.border = thin_border

for _, row in open_items.iterrows():
    ws2.append([
        row["Doc_Number"],
        row["Customer_Name"],
        row["Due_Date"].strftime("%d-%b-%Y"),
        row["Outstanding_Amount"],
        int(row["Days_Overdue"]),
        row["Aging_Bucket"]
    ])
    r = ws2.max_row
    bucket = row["Aging_Bucket"]
    for c in range(1, 7):
        cell = ws2.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = bucket_fills.get(bucket, PatternFill())
        if c == 4:
            cell.number_format = '#,##0.00'
            cell.alignment = right_align

# ── Sheet 3: Collection Efficiency ────────────────────────────
ws3 = wb.create_sheet("Collection Efficiency")

ws3.merge_cells("A1:C1")
ws3["A1"] = "Collection Efficiency Summary"
ws3["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
ws3["A1"].alignment = center_align
ws3.append([])

metrics = [
    ["Total Invoiced (INR)",   total_invoiced],
    ["Total Collected (INR)",  total_collected],
    ["Total Outstanding (INR)",open_items["Outstanding_Amount"].sum()],
    ["Collection Rate (%)",    round(collection_rate, 2)],
]
ws3.append(["Metric", "Value"])
for c in [1, 2]:
    cell = ws3.cell(row=3, column=c)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = center_align
    cell.border = thin_border

for m in metrics:
    ws3.append(m)
    r = ws3.max_row
    ws3.cell(r, 1).font = normal_font
    ws3.cell(r, 1).border = thin_border
    ws3.cell(r, 2).font = normal_font
    ws3.cell(r, 2).border = thin_border
    ws3.cell(r, 2).alignment = right_align
    if "INR" in m[0]:
        ws3.cell(r, 2).number_format = '#,##0.00'

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 20

col_widths2 = [10, 28, 15, 20, 15, 15]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save("reports/ar_aging_report.xlsx")
print("    ✔ Saved → reports/ar_aging_report.xlsx")
print("\n" + "=" * 60)
print("  AR PROCESSING SIMULATION COMPLETE")
print("=" * 60)
