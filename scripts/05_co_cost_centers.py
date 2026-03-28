# SAP CO - Cost Center Accounting (CO-CCA)
# Simulates: KSB1 (Cost Center Line Items), S_ALR_87013611 (Actual vs Plan Report)
# Controlling Area: CA01 | Company Code: IN01 | Fiscal Year: 2024 | Period: 1

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 60)
print("  SAP CO - COST CENTER ACCOUNTING REPORT")
print("  Controlling Area: CA01 | Period: April 2024")
print("=" * 60)

# ── Load Data ──────────────────────────────────────────────────
co = pd.read_csv("transactions/cost_allocations.csv")
cc = pd.read_csv("master_data/cost_centers.csv")
co["Posting_Date"] = pd.to_datetime(co["Posting_Date"])

# ── Step 1: Cost Center Line Items (KSB1) ─────────────────────
print("\n[1] Cost Center Line Items (T-code: KSB1)...")
print(co[["Doc_Number","Posting_Date","Cost_Center","Cost_Center_Name",
          "Description","Actual_Amount","Plan_Amount"]].to_string(index=False))

# ── Step 2: Actual vs Plan by Cost Center (S_ALR_87013611) ────
print("\n[2] Actual vs Plan Report (T-code: S_ALR_87013611)...")
variance = co.groupby(["Cost_Center","Cost_Center_Name"]).agg(
    Total_Actual=("Actual_Amount","sum"),
    Total_Plan=("Plan_Amount","sum")
).reset_index()
variance["Variance"]   = variance["Total_Actual"] - variance["Total_Plan"]
variance["Variance_%"] = ((variance["Variance"] / variance["Total_Plan"]) * 100).round(2)

print(variance.to_string(index=False))
print(f"\n    Grand Total Actual : INR {variance['Total_Actual'].sum():,.2f}")
print(f"    Grand Total Plan   : INR {variance['Total_Plan'].sum():,.2f}")
print(f"    Total Variance     : INR {variance['Variance'].sum():,.2f}")

# ── Step 3: Expense Breakdown by GL Account per Cost Center ───
print("\n[3] Expense Breakdown by GL Account per Cost Center...")
breakdown = co.groupby(["Cost_Center_Name","Description"]).agg(
    Actual=("Actual_Amount","sum"),
    Plan=("Plan_Amount","sum")
).reset_index()
breakdown["Variance"] = breakdown["Actual"] - breakdown["Plan"]
print(breakdown.to_string(index=False))

# ── Step 4: Export to Excel ────────────────────────────────────
print("\n[4] Exporting CO Report to Excel...")

wb = Workbook()

header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
normal_font  = Font(name="Calibri", size=10)
total_font   = Font(name="Calibri", bold=True, size=11)
center_align = Alignment(horizontal="center")
right_align  = Alignment(horizontal="right")
thin_border  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

fav_fill   = PatternFill("solid", fgColor="E2EFDA")
unfav_fill = PatternFill("solid", fgColor="FCE4D6")
total_fill = PatternFill("solid", fgColor="FFF2CC")

# ── Sheet 1: Actual vs Plan ────────────────────────────────────
ws1 = wb.active
ws1.title = "Actual vs Plan (S_ALR)"

ws1.merge_cells("A1:F1")
ws1["A1"] = "PrecisionParts Pvt. Ltd. — Cost Center Actual vs Plan Report"
ws1["A1"].font = Font(name="Calibri", bold=True, size=13, color="375623")
ws1["A1"].alignment = center_align

ws1.merge_cells("A2:F2")
ws1["A2"] = f"Controlling Area: CA01 | Period: April 2024 | T-Code: S_ALR_87013611"
ws1["A2"].font = Font(name="Calibri", size=10, italic=True)
ws1["A2"].alignment = center_align
ws1.append([])

headers = ["Cost Center","Cost Center Name","Total Actual (INR)",
           "Total Plan (INR)","Variance (INR)","Variance (%)"]
ws1.append(headers)
for c in range(1, 7):
    cell = ws1.cell(row=4, column=c)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="375623")
    cell.alignment = center_align
    cell.border = thin_border

for _, row in variance.iterrows():
    is_unfav = row["Variance"] > 0
    row_fill = unfav_fill if is_unfav else fav_fill
    ws1.append([
        row["Cost_Center"],
        row["Cost_Center_Name"],
        row["Total_Actual"],
        row["Total_Plan"],
        row["Variance"],
        f"{row['Variance_%']}%"
    ])
    r = ws1.max_row
    for c in range(1, 7):
        cell = ws1.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = row_fill
        if c in [3, 4, 5]:
            cell.number_format = '#,##0.00'
            cell.alignment = right_align

total_row = ws1.max_row + 1
ws1.append([
    "", "GRAND TOTAL",
    variance["Total_Actual"].sum(),
    variance["Total_Plan"].sum(),
    variance["Variance"].sum(),
    f"{((variance['Variance'].sum()/variance['Total_Plan'].sum())*100):.2f}%"
])
for c in range(1, 7):
    cell = ws1.cell(row=total_row, column=c)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    if c in [3, 4, 5]:
        cell.number_format = '#,##0.00'
        cell.alignment = right_align

ws1.append([])
ws1.append(["", "🟢 Green = Favourable (Actual < Plan)  |  🔴 Red = Unfavourable (Actual > Plan)"])
ws1.cell(ws1.max_row, 2).font = Font(name="Calibri", size=9, italic=True, color="595959")

col_widths = [14, 24, 22, 20, 18, 14]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: Expense Breakdown ────────────────────────────────
ws2 = wb.create_sheet("Expense Breakdown (KSB1)")

ws2.merge_cells("A1:E1")
ws2["A1"] = "PrecisionParts Pvt. Ltd. — Expense Breakdown by Cost Center"
ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color="375623")
ws2["A1"].alignment = center_align

ws2.merge_cells("A2:E2")
ws2["A2"] = f"Controlling Area: CA01 | Period: April 2024 | T-Code: KSB1"
ws2["A2"].font = Font(name="Calibri", size=10, italic=True)
ws2["A2"].alignment = center_align
ws2.append([])

headers2 = ["Cost Center Name","GL Description","Actual (INR)","Plan (INR)","Variance (INR)"]
ws2.append(headers2)
for c in range(1, 6):
    cell = ws2.cell(row=4, column=c)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="375623")
    cell.alignment = center_align
    cell.border = thin_border

for _, row in breakdown.iterrows():
    is_unfav = row["Variance"] > 0
    ws2.append([
        row["Cost_Center_Name"],
        row["Description"],
        row["Actual"],
        row["Plan"],
        row["Variance"]
    ])
    r = ws2.max_row
    for c in range(1, 6):
        cell = ws2.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = unfav_fill if is_unfav else fav_fill
        if c in [3, 4, 5]:
            cell.number_format = '#,##0.00'
            cell.alignment = right_align

col_widths2 = [26, 28, 18, 16, 18]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save("reports/cost_center_variance.xlsx")
print("    ✔ Saved → reports/cost_center_variance.xlsx")
print("\n" + "=" * 60)
print("  CO-CCA SIMULATION COMPLETE")
print("=" * 60)
