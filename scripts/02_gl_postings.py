# SAP FI - General Ledger Postings & Trial Balance
# Simulates: FB50 (GL Posting), FS10N (GL Balance), S_ALR_87012284 (Trial Balance)
# Company Code: IN01 | Fiscal Year: 2024 | Period: 1 (April 2024)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 60)
print("  SAP FI - GL POSTING & TRIAL BALANCE REPORT")
print("  Company Code: IN01 | Period: April 2024")
print("=" * 60)

# ── Load Data ──────────────────────────────────────────────────
gl = pd.read_csv("transactions/gl_journal_entries.csv")
gl["Posting_Date"] = pd.to_datetime(gl["Posting_Date"])

# ── Step 1: Validate Double Entry (Debit = Credit per Document) ─
print("\n[1] Validating double-entry principle (SAP posting check)...")
validation = gl.groupby("Doc_Number").agg(
    Total_Debit=("Debit_Amount", "sum"),
    Total_Credit=("Credit_Amount", "sum")
).reset_index()
validation["Balanced"] = validation["Total_Debit"] == validation["Total_Credit"]

all_balanced = validation["Balanced"].all()
if all_balanced:
    print("    ✔ All documents balanced — Debits = Credits")
else:
    unbalanced = validation[~validation["Balanced"]]
    print(f"    ✘ Unbalanced documents found:\n{unbalanced}")

# ── Step 2: GL Account Summary (FS10N equivalent) ──────────────
print("\n[2] Generating GL Account Balances (T-code: FS10N)...")
gl_summary = gl.groupby(["GL_Account", "Account_Description"]).agg(
    Total_Debit=("Debit_Amount", "sum"),
    Total_Credit=("Credit_Amount", "sum")
).reset_index()
gl_summary["Net_Balance"] = gl_summary["Total_Debit"] - gl_summary["Total_Credit"]
gl_summary = gl_summary.sort_values("GL_Account")
print(gl_summary.to_string(index=False))

# ── Step 3: Trial Balance ──────────────────────────────────────
print("\n[3] Generating Trial Balance (T-code: S_ALR_87012284)...")

coa = pd.read_csv("master_data/chart_of_accounts.csv")
trial_balance = pd.merge(gl_summary, coa[["GL_Account","Account_Type","Account_Group"]],
                         on="GL_Account", how="left")

total_debit  = trial_balance["Total_Debit"].sum()
total_credit = trial_balance["Total_Credit"].sum()
print(f"\n    Total Debits  : INR {total_debit:,.2f}")
print(f"    Total Credits : INR {total_credit:,.2f}")
print(f"    Difference    : INR {(total_debit - total_credit):,.2f}")
if total_debit == total_credit:
    print("    ✔ Trial Balance is BALANCED")
else:
    print("    ✘ Trial Balance is NOT balanced — check entries")

# ── Step 4: Export to Excel ────────────────────────────────────
print("\n[4] Exporting Trial Balance to Excel...")

wb = Workbook()
ws = wb.active
ws.title = "Trial Balance"

# Styles
header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill   = PatternFill("solid", fgColor="1F4E79")
sub_font      = Font(name="Calibri", bold=True, size=10)
sub_fill      = PatternFill("solid", fgColor="BDD7EE")
normal_font   = Font(name="Calibri", size=10)
total_font    = Font(name="Calibri", bold=True, size=11)
total_fill    = PatternFill("solid", fgColor="D6E4F0")
center_align  = Alignment(horizontal="center", vertical="center")
right_align   = Alignment(horizontal="right")
thin_border   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

# Title Block
ws.merge_cells("A1:F1")
ws["A1"] = "PrecisionParts Pvt. Ltd. — Company Code: IN01"
ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
ws["A1"].alignment = center_align

ws.merge_cells("A2:F2")
ws["A2"] = "TRIAL BALANCE | Fiscal Year: 2024 | Period 1 (April 2024)"
ws["A2"].font = Font(name="Calibri", size=11, italic=True)
ws["A2"].alignment = center_align

ws.merge_cells("A3:F3")
ws["A3"] = f"T-Code: S_ALR_87012284 | Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
ws["A3"].font = Font(name="Calibri", size=9, color="808080")
ws["A3"].alignment = center_align

ws.append([])

# Column Headers
headers = ["GL Account", "Account Description", "Account Type",
           "Account Group", "Total Debit (INR)", "Total Credit (INR)"]
ws.append(headers)
for col_num, _ in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Data Rows
for _, row in trial_balance.iterrows():
    ws.append([
        row["GL_Account"],
        row["Account_Description"],
        row["Account_Type"],
        row["Account_Group"],
        row["Total_Debit"],
        row["Total_Credit"]
    ])
    r = ws.max_row
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        cell.font = normal_font
        cell.border = thin_border
        if c in [5, 6]:
            cell.number_format = '#,##0.00'
            cell.alignment = right_align

# Totals Row
ws.append([])
total_row = ws.max_row + 1
ws.append(["", "TOTAL", "", "", total_debit, total_credit])
for c in range(1, 7):
    cell = ws.cell(row=total_row, column=c)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    if c in [5, 6]:
        cell.number_format = '#,##0.00'
        cell.alignment = right_align

# Balanced check row
status = "✔ BALANCED" if total_debit == total_credit else "✘ NOT BALANCED"
ws.append(["", f"Status: {status}", "", "", "", ""])
ws.cell(ws.max_row, 2).font = Font(bold=True, color="375623" if "✔" in status else "C00000")

# Column widths
col_widths = [14, 30, 18, 22, 20, 20]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Save
wb.save("reports/trial_balance.xlsx")
print("    ✔ Saved → reports/trial_balance.xlsx")
print("\n" + "=" * 60)
print("  GL POSTING SIMULATION COMPLETE")
print("=" * 60)
